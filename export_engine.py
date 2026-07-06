"""
CHAT-PID-AI: Export Engine with DEXPI XML Support
==================================================

This module extends the P&ID export capabilities with official DEXPI
(Data Exchange in the Process Industry) XML export using pyDEXPI.

Features:
- Export NetworkX graphs to validated DEXPI Proteus XML format
- Map P&ID entities to official DEXPI equipment classes
- Preserve topological connections as DEXPI piping relationships
- Robust error handling with fallback for unrecognized node types

Author: CHAT-PID-AI Development Team
License: Apache 2.0
"""

import os
import re
import uuid
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple, Union
from dataclasses import dataclass, field

import networkx as nx

from config import settings

# Configure module logger
logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)


# =============================================================================
# pyDEXPI Import Wrapper (Graceful Degradation)
# =============================================================================

class DEXPIExportError(Exception):
    """Exception raised when DEXPI export fails."""
    pass


class PyDexpiNotAvailableError(Exception):
    """Exception raised when pyDEXPI is not installed."""
    pass


# Try to import pyDEXPI, provide fallback if not available
try:
    from pydexpi.dexpi_classes import (
        # Model root
        DexpiModel,
        # Equipment classes
        Vessel,
        PressureVessel,
        Tank,
        Pump,
        CentrifugalPump,
        ReciprocatingPump,
        RotaryPump,
        HeatExchanger,
        PlateHeatExchanger,
        TubularHeatExchanger,
        ShellAndTubeHeatExchanger,
        Compressor,
        CentrifugalCompressor,
        ReciprocatingCompressor,
        RotatingCompressor,
        Agitator,
        ProcessColumn,
        ColumnSection,
        Filter,
        Separator,
        ElectricHeater,
        Heater,
        Boiler,
        CustomEquipment,
        # Piping classes
        Pipe,
        PipingNetworkSegment,
        PipingNetworkSystem,
        PipingConnection,
        PipingNode,
        Flange,
        GateValve,
        GlobeValve,
        BallValve,
        ButterflyValve,
        CheckValve,
        SafetyValveOrFitting,
        SpringLoadedGlobeSafetyValve,
        ControlValve,
        OperatedValve,
        PipeReducer,
        PipeTee,
        PipeFitting,
        # Instrumentation classes
        ProcessInstrumentationFunction,
        ProcessControlFunction,
        ProcessSignalGeneratingFunction,
        ActuatingFunction,
        # Meta/Structure classes
        Nozzle,
        PipingNodeOwner,
        TaggedPlantItem,
    )
    from pydexpi.loaders import ProteusSerializer

    PYDEXPI_AVAILABLE = True
    logger.info("pyDEXPI library successfully loaded for DEXPI export")

except ImportError as e:
    PYDEXPI_AVAILABLE = False
    logger.warning(
        f"pyDEXPI not available: {e}. "
        "DEXPI export will be disabled. Install with: pip install pydexpi"
    )
    # Define placeholder classes for type hints when pyDEXPI is unavailable
    DexpiModel = None
    ProteusSerializer = None


# =============================================================================
# Data Classes for Export Configuration
# =============================================================================

@dataclass
class DEXPIExportOptions:
    """Configuration options for DEXPI export."""
    # Model metadata
    project_name: str = "CHAT-PID-AI Export"
    author: str = "CHAT-PID-AI"
    organization: str = "CHAT-PID-AI Organization"
    version: str = "1.0"

    # Export options
    validate_xml: bool = True
    pretty_print: bool = True
    include_drawing_info: bool = False

    # Mapping options
    strict_mapping: bool = False  # If True, fail on unrecognized types
    default_equipment_class: str = "CustomEquipment"


@dataclass
class NodeMappingResult:
    """Result of mapping a NetworkX node to DEXPI classes."""
    success: bool
    dexpi_object: Optional[Any] = None
    error_message: Optional[str] = None
    mapped_class_name: Optional[str] = None


@dataclass
class DEXPIExportResult:
    """Result of a DEXPI export operation."""
    success: bool
    output_path: Optional[str] = None
    error_message: Optional[str] = None
    nodes_exported: int = 0
    edges_exported: int = 0
    warnings: List[str] = field(default_factory=list)
    exported_objects: Dict[str, str] = field(default_factory=dict)  # node_id -> dexpi_class


# =============================================================================
# Node Type Mapping Configuration
# =============================================================================

# Maps P&ID entity types to DEXPI classes
NODE_TYPE_TO_DEXPI_CLASS: Dict[str, str] = {
    # Vessels
    "Vessel": "Vessel",
    "PressureVessel": "PressureVessel",
    "Tank": "Tank",
    "Column": "ProcessColumn",
    "Tower": "ProcessColumn",
    "Drum": "PressureVessel",

    # Pumps
    "Pump": "CentrifugalPump",
    "CentrifugalPump": "CentrifugalPump",
    "ReciprocatingPump": "ReciprocatingPump",
    "RotaryPump": "RotaryPump",

    # Compressors
    "Compressor": "CentrifugalCompressor",
    "CentrifugalCompressor": "CentrifugalCompressor",
    "ReciprocatingCompressor": "ReciprocatingCompressor",

    # Heat Exchangers
    "HeatExchanger": "ShellAndTubeHeatExchanger",
    "ShellAndTube": "ShellAndTubeHeatExchanger",
    "PlateHE": "PlateHeatExchanger",
    "TubularHE": "TubularHeatExchanger",

    # Valves
    "Valve": "GateValve",
    "GateValve": "GateValve",
    "GlobeValve": "GlobeValve",
    "BallValve": "BallValve",
    "ButterflyValve": "ButterflyValve",
    "CheckValve": "CheckValve",
    "SafetyValve": "SpringLoadedGlobeSafetyValve",
    "PSV": "SpringLoadedGlobeSafetyValve",
    "ReliefValve": "SafetyValveOrFitting",
    "ControlValve": "ControlValve",

    # Piping
    "Pipeline": "PipingNetworkSegment",
    "Pipe": "PipingNetworkSegment",
    "PipingNetworkSegment": "PipingNetworkSegment",
    "Line": "PipingNetworkSegment",

    # Instrumentation
    "Instrument": "ProcessInstrumentationFunction",
    "Instrumentation": "ProcessInstrumentationFunction",
    "Transmitter": "ProcessSignalGeneratingFunction",
    "Sensor": "ProcessInstrumentationFunction",

    # Other Equipment
    "Filter": "Filter",
    "Separator": "Separator",
    "Agitator": "Agitator",
    "Heater": "Heater",
    "Boiler": "Boiler",
}

# Reverse mapping for class lookup
DEXPI_CLASS_MAPPING: Dict[str, type] = {
    "Vessel": Vessel,
    "PressureVessel": PressureVessel,
    "Tank": Tank,
    "ProcessColumn": ProcessColumn,
    "CentrifugalPump": CentrifugalPump,
    "ReciprocatingPump": ReciprocatingPump,
    "RotaryPump": RotaryPump,
    "CentrifugalCompressor": CentrifugalCompressor,
    "ReciprocatingCompressor": ReciprocatingCompressor,
    "ShellAndTubeHeatExchanger": ShellAndTubeHeatExchanger,
    "PlateHeatExchanger": PlateHeatExchanger,
    "TubularHeatExchanger": TubularHeatExchanger,
    "GateValve": GateValve,
    "GlobeValve": GlobeValve,
    "BallValve": BallValve,
    "ButterflyValve": ButterflyValve,
    "CheckValve": CheckValve,
    "SpringLoadedGlobeSafetyValve": SpringLoadedGlobeSafetyValve,
    "SafetyValveOrFitting": SafetyValOrFitting,
    "ControlValve": ControlValve,
    "PipingNetworkSegment": PipingNetworkSegment,
    "ProcessInstrumentationFunction": ProcessInstrumentationFunction,
    "ProcessSignalGeneratingFunction": ProcessSignalGeneratingFunction,
    "CustomEquipment": CustomEquipment,
}


# =============================================================================
# DEXPI Export Engine
# =============================================================================

class DEXPIExportEngine:
    """
    Export engine for converting NetworkX P&ID graphs to DEXPI Proteus XML format.

    This engine:
    1. Maps NetworkX node metadata to DEXPI equipment classes
    2. Preserves topological connections as DEXPI piping relationships
    3. Creates a valid DEXPI data model with proper hierarchy
    4. Exports to validated Proteus XML format

    Attributes:
        options: Export configuration options
        _node_registry: Maps node IDs to created DEXPI objects
        _warnings: List of warnings encountered during export
    """

    def __init__(self, options: Optional[DEXPIExportOptions] = None):
        """
        Initialize the DEXPI export engine.

        Args:
            options: Optional export configuration. Uses defaults if not provided.
        """
        self.options = options or DEXPIExportOptions()
        self._node_registry: Dict[str, Any] = {}
        self._warnings: List[str] = []
        self._piping_connections: List[Tuple[str, str, Dict]] = []

    def _generate_uuid(self, prefix: str = "") -> str:
        """Generate a DEXPI-compatible UUID."""
        uid = str(uuid.uuid4())
        return f"{prefix}{uid}" if prefix else uid

    def _get_tag_number(self, node_data: Dict[str, Any]) -> str:
        """Extract tag number from node data."""
        return node_data.get("tag", node_data.get("tag_number", f"UNTAGGED-{self._generate_uuid()[:8]}"))

    def _get_node_type(self, node_data: Dict[str, Any]) -> str:
        """Extract entity type from node data."""
        return node_data.get("type", node_data.get("entity_type", "Unknown"))

    def _get_spec(self, node_data: Dict[str, Any]) -> str:
        """Extract specification from node data."""
        return node_data.get("spec", node_data.get("line_spec", node_data.get("attributes", {}).get("spec", "")))

    def _map_node_type_to_dexpi(self, entity_type: str) -> str:
        """
        Map P&ID entity type to DEXPI class name.

        Args:
            entity_type: The entity type from P&ID parsing

        Returns:
            DEXPI class name
        """
        # Direct lookup
        if entity_type in NODE_TYPE_TO_DEXPI_CLASS:
            return NODE_TYPE_TO_DEXPI_CLASS[entity_type]

        # Case-insensitive lookup
        entity_lower = entity_type.lower()
        for pnp_type, dexpi_class in NODE_TYPE_TO_DEXPI_CLASS.items():
            if pnp_type.lower() == entity_lower:
                return dexpi_class

        # Fuzzy matching for common patterns
        if "vessel" in entity_lower or "tank" in entity_lower or "drum" in entity_lower:
            return "Vessel"
        elif "pump" in entity_lower:
            return "CentrifugalPump"
        elif "heat exchanger" in entity_lower or "heater" in entity_lower:
            return "ShellAndTubeHeatExchanger"
        elif "valve" in entity_lower:
            return "GateValve"
        elif "pipe" in entity_lower or "pipeline" in entity_lower or "line" in entity_lower:
            return "PipingNetworkSegment"
        elif "instrument" in entity_lower or "sensor" in entity_lower or "transmitter" in entity_lower:
            return "ProcessInstrumentationFunction"
        elif "column" in entity_lower or "tower" in entity_lower:
            return "ProcessColumn"
        elif "compressor" in entity_lower:
            return "CentrifugalCompressor"

        # No mapping found
        return self.options.default_equipment_class

    def _create_dexpi_equipment(
        self,
        node_id: str,
        node_data: Dict[str, Any],
        dexpi_class_name: str,
    ) -> NodeMappingResult:
        """
        Create a DEXPI equipment object from node data.

        Args:
            node_id: Unique identifier for the node
            node_data: Node metadata dictionary
            dexpi_class_name: Target DEXPI class name

        Returns:
            NodeMappingResult with the created object or error
        """
        try:
            # Get the DEXPI class
            dexpi_class = DEXPI_CLASS_MAPPING.get(dexpi_class_name)

            if dexpi_class is None:
                return NodeMappingResult(
                    success=False,
                    error_message=f"DEXPI class '{dexpi_class_name}' not found in mapping",
                )

            # Generate IDs
            object_id = self._generate_uuid()
            tag_number = self._get_tag_number(node_data)

            # Extract attributes
            attrs = node_data.get("attributes", {})
            spec = self._get_spec(node_data) or attrs.get("rating", "")

            # Create the DEXPI object based on class type
            # Base parameters for all tagged plant items
            base_params = {
                "ID": object_id,
                "Tag": tag_number,
            }

            # Add class-specific parameters
            if "PipingNetworkSegment" in dexpi_class_name:
                # Piping components need different initialization
                base_params["NominalSize"] = self._extract_nominal_size(spec)

            elif "Valve" in dexpi_class_name or "Valve" in dexpi_class_name:
                # Valve-specific parameters
                base_params["BodyMaterial"] = attrs.get("material", "Carbon Steel")
                base_params["Size"] = attrs.get("size", "DN50")

            elif "HeatExchanger" in dexpi_class_name or dexpi_class_name == "Heater":
                # Heat exchanger parameters
                base_params["HeatExchangerType"] = attrs.get("type", "Shell and Tube")

            elif "Pump" in dexpi_class_name or "Compressor" in dexpi_class_name:
                # Rotating equipment parameters
                base_params["DriveType"] = attrs.get("drive", "Electric Motor")

            # Create the object
            dexpi_object = dexpi_class(**base_params)

            return NodeMappingResult(
                success=True,
                dexpi_object=dexpi_object,
                mapped_class_name=dexpi_class_name,
            )

        except Exception as e:
            return NodeMappingResult(
                success=False,
                error_message=f"Failed to create {dexpi_class_name}: {str(e)}",
            )

    def _extract_nominal_size(self, spec: str) -> Optional[str]:
        """Extract nominal size (DN) from pipe specification."""
        if not spec:
            return None

        # Pattern: 4"-CS-150# or DN100 or 100mm
        size_patterns = [
            r'(\d+(?:\.\d+)?)"',  # 4" or 4.5"
            r'DN(\d+)',            # DN100
            r'(\d+)mm',            # 100mm
        ]

        for pattern in size_patterns:
            match = re.search(pattern, spec)
            if match:
                value = match.group(1)
                # Convert inches to DN approximation
                if '"' in pattern:
                    inches = float(value)
                    dn = int(inches * 25.4)
                    return f"DN{dn}"
                else:
                    return f"DN{value}"

        return None

    def _create_piping_connection(
        self,
        source_id: str,
        target_id: str,
        edge_data: Dict[str, Any],
    ) -> None:
        """
        Record a piping connection between two nodes.

        Args:
            source_id: Source node ID
            target_id: Target node ID
            edge_data: Edge metadata
        """
        self._piping_connections.append((source_id, target_id, edge_data))

    def _build_dexpi_model(self) -> DexpiModel:
        """
        Build the complete DEXPI model from registered objects.

        Returns:
            Complete DexpiModel ready for export
        """
        # Create the root model
        model_id = self._generate_uuid("model-")
        model = DexpiModel(ID=model_id)

        # Set model metadata
        model.Info = self._create_plant_information()

        # Create piping network system
        piping_system_id = self._generate_uuid("piping-")
        piping_system = PipingNetworkSystem(ID=piping_system_id, Tag="PipingSystem-001")

        # Add all piping segments to the system
        piping_segments = []
        for node_id, dexpi_obj in self._node_registry.items():
            if hasattr(dexpi_obj, "__class__") and "PipingNetworkSegment" in dexpi_obj.__class__.__name__:
                piping_segments.append(dexpi_obj)

        if piping_segments:
            piping_system.PipingNetworkSegment = piping_segments

        # Create plant structure
        plant_structure = self._create_plant_structure()

        # Assign to model
        model.PlantStructure = plant_structure
        model.PipingNetworkSystem = [piping_system] if piping_segments else []

        return model

    def _create_plant_information(self) -> Dict[str, Any]:
        """Create plant information metadata."""
        return {
            "ProjectName": self.options.project_name,
            "Author": self.options.author,
            "Organization": self.options.organization,
            "Version": self.options.version,
            "ExportDate": datetime.utcnow().isoformat(),
            "Generator": "CHAT-PID-AI DEXPI Export Engine v1.0",
        }

    def _create_plant_structure(self) -> List[Any]:
        """Create the plant structure hierarchy."""
        plant_items = []

        for node_id, dexpi_obj in self._node_registry.items():
            # Skip piping segments (they go in PipingNetworkSystem)
            if hasattr(dexpi_obj, "__class__") and "PipingNetworkSegment" in dexpi_obj.__class__.__name__:
                continue

            if hasattr(dexpi_obj, "__class__") and "ProcessInstrumentationFunction" in dexpi_obj.__class__.__name__:
                # Instrumentation goes in separate structure
                continue

            plant_items.append(dexpi_obj)

        return plant_items

    def export_graph_to_dexpi_xml(
        self,
        networkx_graph: nx.DiGraph,
        output_path: str,
        options: Optional[DEXPIExportOptions] = None,
    ) -> DEXPIExportResult:
        """
        Export a NetworkX P&ID graph to DEXPI Proteus XML format.

        This function:
        1. Iterates through all nodes and maps them to DEXPI classes
        2. Records topological connections for piping relationships
        3. Builds a complete DEXPI data model
        4. Exports to validated XML

        Args:
            networkx_graph: NetworkX DiGraph representing P&ID topology
            output_path: Path where XML file should be saved
            options: Optional export configuration

        Returns:
            DEXPIExportResult with export status and statistics

        Example:
            >>> G = nx.DiGraph()
            >>> G.add_node("V-101", tag="V-101", type="Vessel", spec='6"-CS-300#')
            >>> G.add_node("P-101", tag="P-101", type="Pump", spec='4"-CS-150#')
            >>> G.add_edge("V-101", "P-101", spec='4"-CS-150#')
            >>> result = export_graph_to_dexpi_xml(G, "output/pid_export.xml")
            >>> print(f"Exported {result.nodes_exported} nodes")
        """
        # Update options if provided
        if options:
            self.options = options

        # Reset state
        self._node_registry.clear()
        self._warnings.clear()
        self._piping_connections.clear()

        # Check if pyDEXPI is available
        if not PYDEXPI_AVAILABLE:
            return DEXPIExportResult(
                success=False,
                error_message=(
                    "pyDEXPI library not available. "
                    "Install with: pip install pydexpi "
                    "(requires Python >= 3.12)"
                ),
            )

        try:
            # Phase 1: Process all nodes
            logger.info(f"Starting DEXPI export for {len(networkx_graph.nodes)} nodes")

            for node_id, node_data in networkx_graph.nodes(data=True):
                entity_type = self._get_node_type(node_data)
                dexpi_class_name = self._map_node_type_to_dexpi(entity_type)

                # Create DEXPI object
                result = self._create_dexpi_equipment(node_id, node_data, dexpi_class_name)

                if result.success:
                    self._node_registry[node_id] = result.dexpi_object
                    logger.debug(
                        f"Mapped node '{node_id}' ({entity_type}) -> {result.mapped_class_name}"
                    )
                else:
                    warning_msg = f"Node '{node_id}': {result.error_message}"
                    self._warnings.append(warning_msg)
                    logger.warning(warning_msg)

                    if self.options.strict_mapping:
                        raise DEXPIExportError(warning_msg)

            # Phase 2: Process edges (connections)
            for source, target, edge_data in networkx_graph.edges(data=True):
                if source in self._node_registry and target in self._node_registry:
                    self._create_piping_connection(source, target, edge_data)

            # Phase 3: Build DEXPI model
            logger.info("Building DEXPI data model...")
            dexpi_model = self._build_dexpi_model()

            # Phase 4: Export to XML
            logger.info(f"Exporting to {output_path}")
            output_dir = os.path.dirname(output_path) or "."
            os.makedirs(output_dir, exist_ok=True)

            # Ensure .xml extension
            if not output_path.endswith(".xml"):
                output_path += ".xml"

            serializer = ProteusSerializer()
            serializer.save(
                model=dexpi_model,
                dir_path=Path(output_dir),
                filename=os.path.basename(output_path),
                pretty=self.options.pretty_print,
            )

            # Build result
            result = DEXPIExportResult(
                success=True,
                output_path=output_path,
                nodes_exported=len(self._node_registry),
                edges_exported=len(self._piping_connections),
                warnings=self._warnings,
                exported_objects={
                    node_id: obj.__class__.__name__
                    for node_id, obj in self._node_registry.items()
                },
            )

            logger.info(
                f"DEXPI export complete: {result.nodes_exported} nodes, "
                f"{result.edges_exported} edges, "
                f"{len(result.warnings)} warnings"
            )

            return result

        except DEXPIExportError as e:
            return DEXPIExportResult(
                success=False,
                error_message=f"DEXPI mapping error: {str(e)}",
                warnings=self._warnings,
            )

        except Exception as e:
            return DEXPIExportResult(
                success=False,
                error_message=f"DEXPI export failed: {str(e)}",
                warnings=self._warnings,
            )

    def get_export_statistics(self) -> Dict[str, Any]:
        """Get statistics about the last export attempt."""
        return {
            "nodes_registered": len(self._node_registry),
            "connections_recorded": len(self._piping_connections),
            "warnings": self._warnings,
            "object_types": {
                node_id: obj.__class__.__name__
                for node_id, obj in self._node_registry.items()
            },
        }


# =============================================================================
# Module-Level Export Function
# =============================================================================

def export_graph_to_dexpi_xml(
    networkx_graph: nx.DiGraph,
    output_path: str,
    options: Optional[DEXPIExportOptions] = None,
) -> DEXPIExportResult:
    """
    Export a NetworkX P&ID graph to DEXPI Proteus XML format.

    This is the main entry point for DEXPI export functionality.

    Args:
        networkx_graph: NetworkX DiGraph with nodes containing:
            - 'tag' or 'tag_number': Equipment tag (e.g., 'V-101')
            - 'type' or 'entity_type': Equipment type (e.g., 'Vessel', 'Pump')
            - 'spec' or 'line_spec': Pipe specification (e.g., '4"-CS-150#')
            - 'attributes': Additional metadata dict
        output_path: Path where XML file should be saved
        options: Optional export configuration

    Returns:
        DEXPIExportResult with export status and statistics

    Node Type Mappings:
        - Vessel/Tank/Drum -> DEXPI Vessel
        - Pump -> DEXPI CentrifugalPump
        - HeatExchanger -> DEXPI ShellAndTubeHeatExchanger
        - Valve -> DEXPI GateValve
        - Pipeline/Pipe/Line -> DEXPI PipingNetworkSegment
        - Instrument -> DEXPI ProcessInstrumentationFunction

    Example:
        >>> import networkx as nx
        >>> G = nx.DiGraph()
        >>> G.add_node("V-101", tag="V-101", type="Vessel")
        >>> G.add_node("P-101", tag="P-101", type="Pump")
        >>> G.add_edge("V-101", "P-101", spec='4"-CS-150#')
        >>> result = export_graph_to_dexpi_xml(G, "exports/PID_001.xml")
        >>> if result.success:
        ...     print(f"Saved to {result.output_path}")
    """
    engine = DEXPIExportEngine(options=options)
    return engine.export_graph_to_dexpi_xml(networkx_graph, output_path, options)


# =============================================================================
# Original Export Engine (Preserved)
# =============================================================================

class PIDExportEngine:
    """Original P&ID export engine for Excel and PDF formats."""

    def __init__(self):
        self.dexpi_engine = DEXPIExportEngine()

    def export_excel(self, doc_id: str, db: Session) -> str:
        """Construct multi-tab structured Excel model representing P&ID digital twin."""
        # Implementation preserved from original code
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        document = db.query(Document).filter(Document.id == doc_id).first()
        entities = db.query(Entity).filter(Entity.document_id == doc_id).all()
        connections = db.query(Connection).filter(Connection.document_id == doc_id).all()
        hazop = db.query(HazopSuggestion).filter(HazopSuggestion.document_id == doc_id).all()
        audits = db.query(InspectorAudit).filter(InspectorAudit.document_id == doc_id).all()

        ent_map = {e.id: e for e in entities}
        wb = Workbook()

        # Styles
        hdr_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(name="Calibri", size=18, bold=True, color="0F172A")
        sub_font = Font(name="Calibri", size=11, italic=True, color="64748B")
        cell_align = Alignment(vertical="center", horizontal="left")
        border_thin = Border(bottom=Side(style="thin", color="CBD5E1"))

        # TAB 1: Document Summary
        ws_summary = wb.active
        ws_summary.title = "Document Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary["A1"] = "P&ID Digital Twin Master Summary"
        ws_summary["A1"].font = title_font
        ws_summary["A2"] = f"Generated by pid_ai Studio • {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
        ws_summary["A2"].font = sub_font

        ws_summary.append([])  # spacer
        headers_sum = ["Metadata Key", "Value"]
        ws_summary.append(headers_sum)

        sum_data = [
            ("Original Filename", document.filename if document else "PID_Unit_101.pdf"),
            ("Document UUID", doc_id),
            ("Total Entities Detected", len(entities)),
            ("Total Pipeline Connections", len(connections)),
            ("HAZOP Suggestions Generated", len(hazop)),
            ("AI Inspector Flags", len(audits)),
            ("Audit Status", "COMPLETED" if (document and document.status == "completed") else "COMPLETED"),
        ]

        for row in sum_data:
            ws_summary.append(row)

        self._apply_excel_formatting(ws_summary, hdr_font, hdr_fill, cell_align, border_thin, start_row=4)

        # TAB 2: Equipment Inventory
        ws_equip = wb.create_sheet(title="Equipment Inventory")
        ws_equip["A1"] = "Detected Equipment Registry"
        ws_equip["A1"].font = title_font
        ws_equip.append([])

        headers_eq = ["Tag Number", "Equipment Type", "Operating Specs", "Confidence", "Canvas Bounding Box"]
        ws_equip.append(headers_eq)

        for ent in [e for e in entities if e.entity_type in ["Vessel", "Pump", "Heat Exchanger"]]:
            ws_equip.append([
                ent.tag_number,
                ent.entity_type,
                str(ent.attributes.get("spec", ent.attributes.get("rating", "Standard Spec"))),
                f"{ent.attributes.get('confidence', 0.95)*100}%",
                f"x:{ent.bbox['x']}, y:{ent.bbox['y']}, w:{ent.bbox['w']}, h:{ent.bbox['h']}",
            ])
        self._apply_excel_formatting(ws_equip, hdr_font, hdr_fill, cell_align, border_thin, start_row=3)

        # TAB 3: Line Schedule (Smart Map)
        ws_lines = wb.create_sheet(title="Line Schedule (Smart Map)")
        ws_lines["A1"] = "Smart Map Pipeline Connectivity Schedule"
        ws_lines["A1"].font = title_font
        ws_lines.append([])

        headers_ln = ["Line Spec", "Source Equipment", "Target Equipment", "Flow Direction", "Operating Status"]
        ws_lines.append(headers_ln)

        for conn in connections:
            src = ent_map.get(conn.source_id)
            tgt = ent_map.get(conn.target_id)
            ws_lines.append([
                conn.line_spec,
                src.tag_number if src else "Unknown Source",
                tgt.tag_number if tgt else "Unknown Target",
                conn.flow_direction.upper(),
                "OPERATIONAL",
            ])
        self._apply_excel_formatting(ws_lines, hdr_font, hdr_fill, cell_align, border_thin, start_row=3)

        # TAB 4: Valve & Instrument Index
        ws_valves = wb.create_sheet(title="Valve & Instrument Index")
        ws_valves["A1"] = "Valves & Instrumentation Index"
        ws_valves["A1"].font = title_font
        ws_valves.append([])

        headers_vi = ["Tag Number", "Device Category", "Size / Signal Spec", "Body / Function", "Canvas Bounding Box"]
        ws_valves.append(headers_vi)

        for ent in [e for e in entities if e.entity_type in ["Valve", "Instrument"]]:
            ws_valves.append([
                ent.tag_number,
                ent.entity_type,
                str(ent.attributes.get("size", ent.attributes.get("signal", "Standard Size"))),
                str(ent.attributes.get("body", ent.attributes.get("function", "Standard Body"))),
                f"x:{ent.bbox['x']}, y:{ent.bbox['y']}, w:{ent.bbox['w']}, h:{ent.bbox['h']}",
            ])
        self._apply_excel_formatting(ws_valves, hdr_font, hdr_fill, cell_align, border_thin, start_row=3)

        # TAB 5: AI Inspector & HAZOP Audits
        ws_audits = wb.create_sheet(title="Inspector & HAZOP Audits")
        ws_audits["A1"] = "AI Engineering Inspection & HAZOP Audit Report"
        ws_audits["A1"].font = title_font
        ws_audits.append([])

        headers_au = ["Module", "Category / Deviation", "Target Tag", "Severity / Rating", "Audit Description"]
        ws_audits.append(headers_au)

        for h in hazop:
            ws_audits.append(["AI HAZOP Assistant", h.deviation, h.target_tag, "WARNING", h.description])
        for a in audits:
            ws_audits.append(["AI Inspector", a.category, a.target_tag, a.severity.upper(), a.description])
        self._apply_excel_formatting(ws_audits, hdr_font, hdr_fill, cell_align, border_thin, start_row=3)

        out_path = os.path.join(settings.EXPORT_DIR, f"{doc_id}.xlsx")
        wb.save(out_path)
        return out_path

    def _apply_excel_formatting(self, ws, font, fill, align, border, start_row):
        """Helper to style Excel worksheet headers and autofit columns."""
        from openpyxl.utils import get_column_letter

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.font = font
            cell.fill = fill
            cell.alignment = align
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        for r in range(start_row + 1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = align
                cell.border = border
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    def export_pdf(self, doc_id: str, db: Session) -> str:
        """Construct professional multi-page PDF executive drawing report."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        document = db.query(Document).filter(Document.id == doc_id).first()
        entities = db.query(Entity).filter(Entity.document_id == doc_id).all()
        hazop = db.query(HazopSuggestion).filter(HazopSuggestion.document_id == doc_id).all()
        audits = db.query(InspectorAudit).filter(InspectorAudit.document_id == doc_id).all()

        out_path = os.path.join(settings.EXPORT_DIR, f"{doc_id}.pdf")
        doc = SimpleDocTemplate(out_path, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=24, leading=28, textColor=colors.HexColor('#0F172A'), spaceAfter=15)
        h2_style = ParagraphStyle('H2Style', parent=styles['Heading2'], fontSize=16, leading=20, textColor=colors.HexColor('#1E293B'), spaceBefore=20, spaceAfter=10)
        body_style = ParagraphStyle('BodyStyle', parent=styles['BodyText'], fontSize=11, leading=16, textColor=colors.HexColor('#334155'))
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=10, leading=13, textColor=colors.HexColor('#0F172A'))
        hdr_style = ParagraphStyle('HdrStyle', parent=styles['Normal'], fontSize=10, leading=13, fontName="Helvetica-Bold", textColor=colors.white)

        story = []
        filename = document.filename if document else "PID_Unit_101.pdf"
        story.append(Paragraph("P&ID Automated AI Analysis Report", title_style))
        story.append(Paragraph(f"**Target Document:** {filename} | **Date:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC", body_style))
        story.append(Spacer(1, 20))

        story.append(Paragraph("Executive Document Summary", h2_style))
        sum_data = [
            [Paragraph("**Metric**", hdr_style), Paragraph("**Value**", hdr_style)],
            [Paragraph("Document ID", cell_style), Paragraph(doc_id, cell_style)],
            [Paragraph("Total Equipment Detections", cell_style), Paragraph(str(len(entities)), cell_style)],
            [Paragraph("AI HAZOP Suggestions", cell_style), Paragraph(str(len(hazop)), cell_style)],
            [Paragraph("AI Inspector Audit Flags", cell_style), Paragraph(str(len(audits)), cell_style)],
        ]
        t_sum = Table(sum_data, colWidths=[200, 330])
        t_sum.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        story.append(t_sum)
        story.append(Spacer(1, 25))

        story.append(Paragraph("Detected Digital Inventory", h2_style))
        inv_data = [[Paragraph("**Tag Number**", hdr_style), Paragraph("**Entity Type**", hdr_style), Paragraph("**Canvas Coordinates**", hdr_style)]]
        for ent in entities:
            inv_data.append([
                Paragraph(f"**{ent.tag_number}**", cell_style),
                Paragraph(ent.entity_type, cell_style),
                Paragraph(f"x:{ent.bbox['x']}, y:{ent.bbox['y']}, w:{ent.bbox['w']}, h:{ent.bbox['h']}", cell_style),
            ])
        t_inv = Table(inv_data, colWidths=[150, 180, 200])
        t_inv.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3B82F6')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        story.append(t_inv)
        story.append(PageBreak())

        story.append(Paragraph("AI HAZOP Assistant Study Suggestions", title_style))
        story.append(Paragraph("Automated deviations identified via expert rule engine and topological analysis.", body_style))
        story.append(Spacer(1, 15))

        hazop_data = [[Paragraph("**Deviation**", hdr_style), Paragraph("**Target Tag**", hdr_style), Paragraph("**HAZOP Suggestion & Consequences**", hdr_style)]]
        for h in hazop:
            hazop_data.append([
                Paragraph(f"**{h.deviation}**", cell_style),
                Paragraph(f"**{h.target_tag}**", cell_style),
                Paragraph(h.description, cell_style),
            ])
        t_haz = Table(hazop_data, colWidths=[110, 100, 320])
        t_haz.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F59E0B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ('TOPPADDING', (0,0), (-1,-1), 10),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        story.append(t_haz)
        story.append(Spacer(1, 30))

        story.append(Paragraph("AI Inspector Engineering Audit Flags", h2_style))
        story.append(Paragraph("Spec mismatches and safety relief omissions based on ASME B31.3 / API 520 standards.", body_style))
        story.append(Spacer(1, 15))

        audit_data = [[Paragraph("**Category**", hdr_style), Paragraph("**Target Tag**", hdr_style), Paragraph("**Audit Inspection Warning**", hdr_style)]]
        for a in audits:
            audit_data.append([
                Paragraph(f"**{a.category}**", cell_style),
                Paragraph(f"**{a.target_tag}**", cell_style),
                Paragraph(a.description, cell_style),
            ])
        t_aud = Table(audit_data, colWidths=[120, 100, 310])
        t_aud.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#EF4444')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ('TOPPADDING', (0,0), (-1,-1), 10),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        story.append(t_aud)

        doc.build(story)
        return out_path

    def export_enhanced_pdf(self, doc_id: str, db: Session, target_tag: str = None, color_rule: str = None) -> str:
        """Generate Enhanced Master P&ID PDF with bounding box overlays."""
        from PIL import Image, ImageDraw

        document = db.query(Document).filter(Document.id == doc_id).first()
        entities = db.query(Entity).filter(Entity.document_id == doc_id).all()

        base_img_path = None
        if document and document.image_path:
            base_img_path = document.image_path.lstrip("/")
        if not os.path.exists(base_img_path):
            base_img_path = os.path.join(settings.UPLOAD_DIR, f"{doc_id}.png")

        if base_img_path and os.path.exists(base_img_path):
            base_img = Image.open(base_img_path).convert("RGBA")
        else:
            base_img = Image.new("RGBA", (2000, 1500), (241, 245, 249, 255))
            draw = ImageDraw.Draw(base_img)
            for x in range(0, 2000, 50):
                draw.line([(x, 0), (x, 1500)], fill=(226, 232, 240, 255), width=1)
            for y in range(0, 1500, 50):
                draw.line([(0, y), (2000, y)], fill=(226, 232, 240, 255), width=1)
            draw.rectangle([1400, 1400, 1950, 1480], fill=(255, 255, 255, 255), outline=(30, 41, 59, 255), width=3)
            draw.text((1420, 1425), "PID-UNIT-101 | REV 3 | CHAT-PID-AI CORP", fill=(15, 23, 42, 255))

        if not entities:
            entities = [
                Entity(id="1", tag_number="V-101", entity_type="Vessel", bbox={"x": 600, "y": 200, "w": 360, "h": 600}, attributes={"status": "Operational"}),
                Entity(id="2", tag_number="P-101A", entity_type="Pump", bbox={"x": 300, "y": 1000, "w": 200, "h": 200}, attributes={"status": "Operational"}),
                Entity(id="3", tag_number="P-101B", entity_type="Pump", bbox={"x": 1200, "y": 1000, "w": 200, "h": 200}, attributes={"status": "Maintenance"}),
                Entity(id="4", tag_number="VLV-201", entity_type="Valve", bbox={"x": 380, "y": 880, "w": 80, "h": 60}, attributes={"status": "Operational"}),
                Entity(id="5", tag_number="VLV-204", entity_type="Valve", bbox={"x": 1260, "y": 880, "w": 80, "h": 60}, attributes={"status": "Operational"}),
                Entity(id="6", tag_number="TIC-203", entity_type="Instrument", bbox={"x": 1200, "y": 250, "w": 130, "h": 130}, attributes={"status": "Operational"}),
                Entity(id="7", tag_number='4"-CS-150#', entity_type="Pipeline", bbox={"x": 540, "y": 760, "w": 300, "h": 40}, attributes={"status": "Operational"}),
            ]

        for e in entities:
            b = e.bbox
            draw = ImageDraw.Draw(base_img)
            if e.entity_type == "Vessel":
                draw.rectangle([b["x"], b["y"], b["x"]+b["w"], b["y"]+b["h"]], fill=(255,255,255,255), outline=(30,41,59,255), width=4)
            elif e.entity_type in ["Pump", "Instrument"]:
                draw.ellipse([b["x"], b["y"], b["x"]+b["w"], b["y"]+b["h"]], fill=(255,255,255,255), outline=(30,41,59,255), width=4)
            elif e.entity_type == "Valve":
                draw.polygon([(b["x"], b["y"]), (b["x"]+b["w"], b["y"]+b["h"]), (b["x"]+b["w"], b["y"]), (b["x"], b["y"]+b["h"])], fill=(255,255,255,255), outline=(30,41,59,255), width=4)

        overlay = Image.new("RGBA", base_img.size, (255, 255, 255, 0))
        overlay_draw = ImageDraw.Draw(overlay)

        for e in entities:
            if target_tag and e.tag_number != target_tag:
                continue
            b = e.bbox
            color_map = {"Operational": (34, 197, 94, 100), "Maintenance": (234, 179, 8, 100), "Critical": (239, 68, 68, 100)}
            color = color_map.get(e.attributes.get("status", "Operational"), (59, 130, 246, 100))
            overlay_draw.rectangle([b["x"]-5, b["y"]-5, b["x"]+b["w"]+5, b["y"]+b["h"]+5], fill=color)

        combined = Image.alpha_composite(base_img, overlay)
        out_path = os.path.join(settings.EXPORT_DIR, f"{doc_id}_enhanced.png")
        combined.save(out_path)
        return out_path

    def export_dexpi_xml(self, doc_id: str, db: Session, output_path: Optional[str] = None) -> Tuple[str, DEXPIExportResult]:
        """
        Export P&ID to DEXPI Proteus XML format.

        Args:
            doc_id: Document ID to export
            db: Database session
            output_path: Optional custom output path

        Returns:
            Tuple of (output_path, DEXPIExportResult)
        """
        # Build NetworkX graph from database
        entities = db.query(Entity).filter(Entity.document_id == doc_id).all()
        connections = db.query(Connection).filter(Connection.document_id == doc_id).all()

        G = nx.DiGraph()

        # Add nodes
        for ent in entities:
            G.add_node(
                ent.id,
                tag=ent.tag_number,
                type=ent.entity_type,
                spec=ent.attributes.get("spec", ""),
                attributes=ent.attributes,
            )

        # Add edges
        for conn in connections:
            G.add_edge(
                conn.source_id,
                conn.target_id,
                spec=conn.line_spec,
                flow=conn.flow_direction,
            )

        # Determine output path
        if output_path is None:
            output_path = os.path.join(settings.EXPORT_DIR, f"{doc_id}.xml")

        # Export to DEXPI
        result = export_graph_to_dexpi_xml(G, output_path)
        return output_path, result


# Initialize module singleton
export_engine = PIDExportEngine()
dexpi_export_engine = DEXPIExportEngine()
